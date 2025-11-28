import streamlit as st
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook

# 1. 페이지 설정 및 제목
st.set_page_config(page_title="바이크 정비노트", page_icon="🏍️", layout="wide")
st.title("🏍️ 마이 바이크 정비노트")

FILE_NAME = '오토바이_정비내역.xlsx'

# --- 탭 구분 (입력하기 / 내역보기) ---
tab1, tab2 = st.tabs(["📝 정비 입력", "📋 전체 내역 조회"])

# ==========================================
# [탭 1] 정비 내용 입력하기
# ==========================================
with tab1:
    st.subheader("새로운 정비 내용 추가")

    # 빠른 저장 함수
    def save_fast(data_list):
        if not os.path.exists(FILE_NAME):
            wb = Workbook()
            ws = wb.active
            ws.append(['날짜', '차종', '주행거리(km)', '항목', '내용', '비용(원)', '기록일시'])
            wb.save(FILE_NAME)
        
        wb = load_workbook(FILE_NAME)
        ws = wb.active
        ws.append(data_list)
        wb.save(FILE_NAME)

    # 입력 폼
    with st.form(key='maintenance_form', clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            date = st.date_input("정비 날짜", datetime.now())
            bike_model = st.text_input("차종", value="존테스 350D")
        with col2:
            mileage = st.number_input("주행 거리 (km)", step=100)
            cost = st.number_input("비용 (원)", step=1000)

        category = st.selectbox("정비 항목", ["엔진오일", "타이어", "브레이크", "구동계", "전기장치", "기타", "주유"])
        details = st.text_area("상세 내용 (예: 합성유 100%, 공임포함)", height=80)
        
        submit_button = st.form_submit_button(label='💾 저장하기', use_container_width=True)

    if submit_button:
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data_to_save = [date, bike_model, mileage, category, details, cost, current_time]
        try:
            save_fast(data_to_save)
            st.success(f"✅ 저장되었습니다! (탭2에서 확인 가능)")
        except Exception as e:
            st.error(f"저장 실패: {e}")

# ==========================================
# [탭 2] 전체 내역 조회하기
# ==========================================
with tab2:
    st.subheader("📋 정비 기록 대장")
    
    if os.path.exists(FILE_NAME):
        # 엑셀 데이터 불러오기
        df = pd.read_excel(FILE_NAME)
        
        if not df.empty:
            # 1. 보기 좋게 가공
            df['날짜'] = pd.to_datetime(df['날짜']).dt.date # 시간 떼고 날짜만
            df = df.sort_values(by='날짜', ascending=False) # 최신순 정렬
            
            # 2. 통계 보여주기 (총 비용, 총 정비 횟수)
            total_cost = df['비용(원)'].sum()
            total_count = len(df)
            
            # 메트릭(지표) 표시
            m_col1, m_col2, m_col3 = st.columns(3)
            m_col1.metric("총 정비 비용", f"{total_cost:,.0f}원")
            m_col2.metric("총 기록 횟수", f"{total_count}회")
            m_col3.metric("최근 정비일", str(df.iloc[0]['날짜']))

            st.divider() # 구분선

            # 3. 데이터 테이블 표시
            # dataframe은 열 클릭시 정렬 가능, 돋보기로 검색 가능
            st.dataframe(df, use_container_width=True, hide_index=True)
            
        else:
            st.info("데이터가 비어있습니다. '정비 입력' 탭에서 내용을 추가해주세요.")
    else:
        st.warning("아직 생성된 정비 기록 파일이 없습니다.")
